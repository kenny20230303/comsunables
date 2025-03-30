# admin_bp
from flask import Blueprint, request, jsonify
from db.db_config import db_session, UserDepartment, User, Supplier, Material, MonthlyPlan, DailyApplication,MaterialHistory
from datetime import datetime
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from flask import send_file, make_response
from io import BytesIO

finance_bp = Blueprint('finance', __name__, url_prefix='/finance')


# todo 月度计划管理
@finance_bp.route('/getMonthlyPlan', methods=['POST', 'PUT'])
def getMonthlyPlan():
    if request.method == 'POST':
        with db_session() as session:
            res = session.query(MonthlyPlan).filter(MonthlyPlan.MonthlyPlan_type == '计划内').all()
            infos_list = []
            for i in res:
                # 获取用户信息
                User_info = db_session.query(User).filter(User.id == int(i.user_id)).first()
                # 获取商品信息
                Material_info = db_session.query(Material).filter(Material.id == int(i.Material_id)).first()
                if i.auditor == 'True':
                    infos_list.append({'name': Material_info.name, 'specifications': Material_info.specifications,
                                       'unit': Material_info.unit, 'price': Material_info.price,
                                       'id': i.id,

                                       'user_name': User_info.name, 'user_phone': User_info.phone,

                                       'Material_number': i.Material_number, 'finance': i.finance,
                                       'number_change': i.number_change, 'current_time': i.current_time

                                       })
        infos_list.reverse()
        return jsonify(infos_list)
    if request.method == 'PUT':
        data = request.get_json()
        material_id = data['id']
        new_number = data['number']

        with db_session() as session:
            monthly_plan = session.query(MonthlyPlan).filter(MonthlyPlan.id == material_id).first()
            if monthly_plan:

                if monthly_plan.Material_number != new_number:
                    monthly_plan.number_change = 'True'

                monthly_plan.Material_number = new_number
                monthly_plan.finance = 'True'
                session.commit()
                return jsonify({'message': 'Update successful'}), 200
            else:
                return jsonify({'message': 'Material not found'}), 404


# todo 月度计划管理
@finance_bp.route('/getMonthlyPlanOuter', methods=['POST', 'PUT'])
def getMonthlyPlanOuter():
    if request.method == 'POST':
        with db_session() as session:
            res = session.query(MonthlyPlan).filter(MonthlyPlan.MonthlyPlan_type == '计划外').all()

            infos_list = []
            for i in res:
                # 获取用户信息
                User_info = db_session.query(User).filter(User.id == int(i.user_id)).first()
                # 获取商品信息
                Material_info = db_session.query(Material).filter(Material.id == int(i.Material_id)).first()
                if i.auditor == 'True':
                    infos_list.append({'name': Material_info.name, 'specifications': Material_info.specifications,
                                       'unit': Material_info.unit, 'price': Material_info.price,
                                       'id': i.id,

                                       'user_name': User_info.name, 'user_phone': User_info.phone,

                                       'Material_number': i.Material_number, 'finance': i.finance,
                                       'number_change': i.number_change, 'current_time': i.current_time

                                       })
        infos_list.reverse()
        return jsonify(infos_list)
    if request.method == 'PUT':
        data = request.get_json()
        material_id = data['id']
        new_number = data['number']

        with db_session() as session:
            monthly_plan = session.query(MonthlyPlan).filter(MonthlyPlan.id == material_id).first()
            if monthly_plan:

                if monthly_plan.Material_number != new_number:
                    monthly_plan.number_change = 'True'
                monthly_plan.MonthlyPlan_type = '计划外'
                monthly_plan.Material_number = new_number
                monthly_plan.finance = 'True'
                session.commit()
                return jsonify({'message': 'Update successful'}), 200
            else:
                return jsonify({'message': 'Material not found'}), 404


# todo 日度计划管理
@finance_bp.route('/getDayPlan', methods=['POST', 'PUT'])
def getDayPlan():
    if request.method == 'POST':
        with db_session() as session:
            res = session.query(DailyApplication).all()
            infos_list = []
            for i in res:
                User_info = db_session.query(User).filter(User.id == int(i.user_id)).first()
                Material_info = db_session.query(Material).filter(Material.id == int(i.Material_id)).first()
                if i.auditor == 'True':
                    infos_list.append({'name': Material_info.name, 'specifications': Material_info.specifications,
                                       'unit': Material_info.unit, 'price': Material_info.price,
                                       'id': i.id,
                                       'Material_info_id': Material_info.id,

                                       'user_name': User_info.name, 'user_phone': User_info.phone,

                                       'Material_number': i.Material_number, 'finance': i.finance,
                                       'number_change': i.number_change, 'current_time': i.current_time

                                       })
        infos_list.reverse()
        return jsonify(infos_list)
    if request.method == 'PUT':
        data = request.get_json()
        Material_info_id = data['Material_info_id']
        DailyApplication_id = data['DailyApplication_id']
        new_number = data['number']

        res = db_session.query(Material).filter(Material.id == int(Material_info_id)).first()
        if int(new_number) > int(res.number):
            return jsonify({'message': '库存数量少'}), 404

        with db_session() as session:
            monthly_plan = session.query(DailyApplication).filter(DailyApplication.id == int(DailyApplication_id)).first()
            if monthly_plan:

                if monthly_plan.Material_number != new_number:
                    monthly_plan.number_change = 'True'

                monthly_plan.Material_number = new_number
                monthly_plan.finance = 'True'

                res.number = str(int(res.number) - int(new_number))
                db_session.commit()

                # 创建历史记录
                history_log = MaterialHistory(
                    name=res.name,
                    specifications=res.specifications,
                    unit=res.unit,
                    price=res.price,
                    Supplier_id=res.Supplier_id,
                    number=str(new_number),  # 记录本次入库数量
                    type='出库',  # 记录类型
                )
                session.add(history_log)  # 添加到会话

                session.commit()

                return jsonify({'message': 'Update successful'}), 200
            else:
                return jsonify({'message': 'Material not found'}), 404


# todo 月度计划管理
@finance_bp.route('/getMonthlyPlanEchars', methods=['POST', 'GET'])
def getMonthlyPlanEchars():
    if request.method == 'GET':
        try:
            with db_session():
                res = db_session.query(MonthlyPlan).all()
                infos_list = []
                for i in res:
                    if isinstance(i.belong_month, str) and len(i.belong_month) >= 7:
                        infos_list.append(i.belong_month[:7])
                unique_infos = list(set(infos_list))
                sorted_infos = sorted(unique_infos, reverse=True)

                sorted_infos = [{'value': i, 'label': i} for i in sorted_infos]
                print(sorted_infos)

                return jsonify(sorted_infos)

        except Exception as e:
            # 处理异常并返回错误信息
            print(f"An error occurred: {e}")
            return jsonify({"error": str(e)}), 500

    if request.method == 'POST':
        current_month = request.get_json()['value']

        # 获取所有材料信息和部门名称
        materials = db_session.query(Material).all()
        departments = db_session.query(UserDepartment.name).all()
        department_names = [dept.name for dept in departments]

        # 准备结果列表
        result = []

        # 遍历每个材料
        for material in materials:
            supplier = db_session.query(Supplier).filter(Supplier.id == int(material.Supplier_id)).first()
            supplier_name = supplier.number if supplier else '未知'
            price = float(material.price)

            material_info = {
                '编号': material.id,
                '名称': material.name,
                '规格': material.specifications,
                '单位': material.unit,
                '供应商名称': supplier_name,
                '价格': price,
                '安全库存数量': int(material.secure_number) * int(material.secure_days),
            }

            total_quantity = 0

            for department_name in department_names:
                quantity = db_session.query(func.sum(MonthlyPlan.Material_number)) \
                               .join(User, User.id == MonthlyPlan.user_id) \
                               .filter(
                    MonthlyPlan.Material_id == material.id,
                    User.UserDepartment_name == department_name,
                    MonthlyPlan.MonthlyPlan_type == '计划内',
                    MonthlyPlan.belong_month.startswith(current_month)
                ).scalar() or 0
                total_quantity += quantity
                material_info[department_name] = quantity

            material_info['本月需求数量'] = total_quantity
            material_info['本月需求金额'] = total_quantity * price
            material_info['预计数量'] = total_quantity + material_info['安全库存数量']
            material_info['预计金额'] = material_info['预计数量'] * price

            result.append(material_info)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "MaterialInformation"

        headers = ['编号', '名称', '规格', '单位', '供应商名称', '价格', '安全库存数量'] + department_names + ['本月需求数量', '本月需求金额', '预计数量', '预计金额', '合计', '第一次', '第二次',
                                                                                                               '第三次']
        sheet.append(headers)

        for item in result:
            row = [item.get(header, '') for header in headers]
            sheet.append(row)

        total_amount = sum(item['本月需求金额'] for item in result)
        total_row = ['' for _ in range(len(headers))]
        total_row[headers.index('本月需求金额')] = f"合计金额: {total_amount}"
        sheet.append(total_row)

        approval_info = ["承认:", " 课长", " 系长申购: "]
        approval_row = approval_info + ['' for _ in range(len(headers) - len(approval_info))]
        sheet.append(approval_row)

        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 10)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # 创建内存中的文件
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # 返回文件作为响应
        response = make_response(
            send_file(output, as_attachment=True, download_name=f"{current_month}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
        response.headers["Content-Disposition"] = f"attachment; filename={current_month}.xlsx"
        return response


# todo 月度计划管理
@finance_bp.route('/getMonthlyPlanOuterEchars', methods=['POST', 'GET'])
def getMonthlyPlanOuterEchars():
    if request.method == 'GET':
        try:
            with db_session():
                res = db_session.query(MonthlyPlan).all()
                infos_list = []
                for i in res:
                    if isinstance(i.belong_month, str) and len(i.belong_month) >= 7:
                        infos_list.append(i.belong_month[:7])
                unique_infos = list(set(infos_list))
                sorted_infos = sorted(unique_infos, reverse=True)

                sorted_infos = [{'value': i, 'label': i} for i in sorted_infos]
                print(sorted_infos)

                return jsonify(sorted_infos)

        except Exception as e:
            # 处理异常并返回错误信息
            print(f"An error occurred: {e}")
            return jsonify({"error": str(e)}), 500

    if request.method == 'POST':
        # 获取当前月份
        current_month = request.get_json()['value']
        # 获取所有材料信息和部门名称
        materials = db_session.query(Material).all()
        departments = db_session.query(UserDepartment.name).all()
        department_names = [dept.name for dept in departments]

        # 准备结果列表
        result = []

        # 遍历每个材料
        for material in materials:
            supplier = db_session.query(Supplier).filter(Supplier.id == int(material.Supplier_id)).first()
            supplier_name = supplier.number if supplier else '未知'
            price = float(material.price)

            # 查询与当前材料相关的MonthlyPlan记录，获取纳期和备注
            monthly_plan = db_session.query(MonthlyPlan).filter(
                MonthlyPlan.Material_id == material.id,
                MonthlyPlan.MonthlyPlan_type == '计划外',
                MonthlyPlan.belong_month.startswith(current_month)
            ).first()

            naqi = monthly_plan.naqi if monthly_plan else ''
            beizhu = monthly_plan.beizhu if monthly_plan else ''

            material_info = {
                '编号': material.id,
                '名称': material.name,
                '规格': material.specifications,
                '单位': material.unit,
                '供应商名称': supplier_name,
                '价格': price,
                '纳期': naqi,
                '备注': beizhu,
            }

            total_quantity = 0

            for department_name in department_names:
                quantity = db_session.query(func.sum(MonthlyPlan.Material_number)) \
                               .join(User, User.id == MonthlyPlan.user_id) \
                               .filter(
                    MonthlyPlan.Material_id == material.id,
                    User.UserDepartment_name == department_name,
                    MonthlyPlan.MonthlyPlan_type == '计划外',
                    MonthlyPlan.belong_month.startswith(current_month)
                ).scalar() or 0
                total_quantity += quantity
                material_info[department_name] = quantity

            material_info['本月需求数量'] = total_quantity
            material_info['本月需求金额'] = total_quantity * price

            result.append(material_info)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "MaterialInformation"

        # 更新标题行，添加纳期和备注
        headers = ['编号', '名称', '规格', '单位', '供应商名称', '价格'] + department_names + ['本月需求数量', '本月需求金额', '纳期', '备注']
        sheet.append(headers)

        for item in result:
            row = [item.get(header, '') for header in headers]
            sheet.append(row)

        total_amount = sum(item['本月需求金额'] for item in result)
        total_row = ['' for _ in range(len(headers))]
        total_row[headers.index('本月需求金额')] = f"合计金额: {total_amount}"
        sheet.append(total_row)

        # 自动调整列宽并设置居中对齐
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 10)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # 创建内存中的文件
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # 返回文件作为响应
        response = make_response(
            send_file(output, as_attachment=True, download_name=f"{current_month}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
        response.headers["Content-Disposition"] = f"attachment; filename={current_month}.xlsx"
        return response


def get_unique_months():
    """获取 DailyApplication 表中唯一的月份列表。"""
    results = db_session.query(
        func.strftime('%Y-%m', DailyApplication.belong_month).label('month')
    ).distinct().all()
    return [result.month for result in results]


def get_daily_consumption_for_month(month):
    """获取指定月份的日消耗物品记录。"""
    daily_consumptions = db_session.query(
        DailyApplication.Material_id,
        Material.name,
        Material.specifications,
        Material.price,
        DailyApplication.Material_number
    ).join(Material, DailyApplication.Material_id == Material.id) \
        .filter(func.strftime('%Y-%m', DailyApplication.belong_month) == month).all()

    results = []
    for item in daily_consumptions:
        total_price = float(item.price) * float(item.Material_number)
        results.append({
            'name': item.name,
            'specifications': item.specifications,
            'unit_price': item.price,
            'quantity': item.Material_number,
            'total_price': total_price
        })
    return results


def get_department_consumption_for_month(month):
    """获取指定月份按部门分组的日消耗物品记录。"""
    department_consumptions = db_session.query(
        UserDepartment.name.label('department_name'),  # 使用 label 来明确表示
        DailyApplication.Material_id,
        Material.name.label('material_name'),  # 使用 label 来明确表示
        Material.specifications,
        Material.price,
        DailyApplication.Material_number
    ).join(User, User.id == DailyApplication.user_id) \
        .join(Material, DailyApplication.Material_id == Material.id) \
        .join(UserDepartment, User.UserDepartment_name == UserDepartment.name) \
        .filter(func.strftime('%Y-%m', DailyApplication.belong_month) == month).all()

    results = []
    for item in department_consumptions:
        total_price = float(item.price) * float(item.Material_number)
        results.append({
            'department_name': item.department_name,  # 使用 label 获取部门名称
            'name': item.material_name,  # 使用 label 获取材料名称
            'specifications': item.specifications,
            'unit_price': item.price,
            'quantity': item.Material_number,
            'total_price': total_price
        })
    return results


# todo 月度计划管理
@finance_bp.route('/getMonthlyPlanHtml', methods=['POST', 'GET'])
def getMonthlyPlanHtml():
    if request.method == 'GET':
        unique_months = get_unique_months()
        res = [{'value': i, 'label': i} for i in unique_months]
        return res

    if request.method == 'POST':
        month = request.get_json()['month']
        first_month = month
        daily_consumption = get_daily_consumption_for_month(first_month)
        department_consumption = get_department_consumption_for_month(first_month)
        print(daily_consumption)
        print(department_consumption)
        return jsonify({'daily_consumption': daily_consumption, 'department_consumption': department_consumption})
